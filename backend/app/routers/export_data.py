import io
import pandas as pd
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from ..database import get_db
from ..models import Customer
from .customers import get_customers_query
from .analytics import get_customer_movement_data
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side

router = APIRouter(prefix="/api/export", tags=["export"])

@router.get("/excel")
async def export_customers_excel(
    db: Session = Depends(get_db),
    search: str = Query(None),
    nhom_kh: str = Query(None),
    rfm_segment: str = Query(None),
    loai_kh: str = Query(None),
    chu_y_churn: bool = Query(False),
    has_debt: bool = Query(False),
    has_revenue: bool = Query(False),
    rev_range: str = Query(None),
    ban_giao_status: str = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    sort_by: str = Query("tong_doanh_thu"),
    order: str = Query("desc")
):
    # Sử dụng logic query tập trung đã refactor (Dùng keyword arguments để tránh lệch filter)
    query, revenue_col = get_customers_query(
        db=db, 
        search=search, 
        nhom_kh=nhom_kh, 
        rfm_segment=rfm_segment, 
        loai_kh=loai_kh, 
        chu_y_churn=chu_y_churn,
        has_debt=has_debt,
        has_revenue=has_revenue,
        rev_range=rev_range, 
        ban_giao_status=ban_giao_status, 
        start_date=start_date, 
        end_date=end_date, 
        sort_by=sort_by, 
        order=order
    )
    
    # Lấy toàn bộ kết quả đã lọc (không phân trang)
    results = query.all()
    
    data = []
    for idx, (c, rev) in enumerate(results):
        # Logic Trạng thái đồng bộ với UI (Dựa trên doanh thu dải ngày đã chọn)
        dynamic_revenue = float(rev) if rev is not None else 0.0
        trang_thai = "Ko phát sinh" if dynamic_revenue <= 0 else "Hoạt động"
        
        data.append({
            "STT": idx + 1,
            "Mã CRM/CMS": c.ma_crm_cms,
            "Tên KH": c.ten_kh,
            "Phân Loại": c.nhom_kh,
            "Phân Khúc": c.rfm_segment,
            "Tình trạng bàn giao": c.tinh_hinh_ban_giao_cms or "Chưa bàn giao",
            "D.Thu (Trong kỳ)": dynamic_revenue,
            "Trạng Thái": trang_thai,
            # Các cột ẩn (Sếp yêu cầu giữ lại)
            "Loại KH": c.loai_kh,
            "Đơn vị (Tên BC/VHX)": c.ten_bc_vhx,
            "Bộ phận (BĐP/X)": c.bdp_x
        })
        
    # Tạo DataFrame với thứ tự cột cố định để tránh lệch
    columns = [
        "STT", "Mã CRM/CMS", "Tên KH", "Phân Loại", "Phân Khúc", 
        "Tình trạng bàn giao", "D.Thu (Trong kỳ)", "Trạng Thái",
        "Loại KH", "Đơn vị (Tên BC/VHX)", "Bộ phận (BĐP/X)"
    ]
    df = pd.DataFrame(data, columns=columns)
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="DanhSachKhachHang")
        
        workbook = writer.book
        worksheet = writer.sheets["DanhSachKhachHang"]
        
        # --- ĐỊNH DẠNG PREMIUM VNPOST ---
        
        # 1. Colors & Styles
        vnpost_blue = "0054A6"
        vnpost_orange = "F9A51A"
        danger_red = "DC2626"
        success_green = "16A34A"
        
        header_fill = PatternFill(start_color=vnpost_blue, end_color=vnpost_blue, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        alternate_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        # 2. Header Row
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # 3. Data Rows
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), 2):
            is_alternate = row_idx % 2 == 0
            trang_thai_raw = row[7].value # Cột Trạng Thái (Index 7)
            
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.font = Font(size=10)
                
                if is_alternate:
                    cell.fill = alternate_fill
                
                # Căn lề và Màu sắc theo cột
                if col_idx in [1, 2, 4, 5, 8]: # STT, Mã, Phân loại, Phân khúc, Trạng thái
                    cell.alignment = center_align
                elif col_idx == 7: # Doanh thu
                    cell.alignment = right_align
                    cell.number_format = '#,##0 "đ"'
                    cell.font = Font(bold=True, size=10)
                else:
                    cell.alignment = left_align
                
                # Highlight đặc biệt cho Trạng Thái
                if col_idx == 8:
                    if trang_thai_raw == "Ko phát sinh":
                        cell.font = Font(color=danger_red, bold=True)
                    else:
                        cell.font = Font(color=success_green, bold=True)
        
        # 4. Tự động chỉnh độ rộng cột
        column_widths = {
            1: 5,   # STT
            2: 15,  # Mã
            3: 35,  # Tên KH
            4: 20,  # Phân Loại
            5: 15,  # Phân Khúc
            6: 25,  # Bàn giao
            7: 18,  # Doanh thu
            8: 15,  # Trạng thái
            9: 25,  # Loại KH
            10: 30, # Đơn vị
            11: 20  # BĐP/X
        }
        for col_idx, width in column_widths.items():
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = width
            
        worksheet.freeze_panes = "A2"

    buffer.seek(0)
    
    filename = "BaoCao_KHHH_Hue_Filtered.xlsx"
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

@router.get("/customer-movement")
async def export_customer_movement_excel(
    db: Session = Depends(get_db),
    current_start_date: str = Query(None),
    current_end_date: str = Query(None),
    compare_start_date: str = Query(None),
    compare_end_date: str = Query(None),
    rfm_segment: str = Query(None),
    movement_status: str = Query(None),
    nhom_kh: str = Query(None),
    nhan_su: str = Query(None),
    search: str = Query(None)
):
    items, _, _, _, _ = get_customer_movement_data(
        current_start_date, current_end_date, compare_start_date, compare_end_date,
        rfm_segment, movement_status, nhom_kh, nhan_su, search, db
    )
    
    data = []
    for idx, c in enumerate(items):
        data.append({
            "STT": idx + 1,
            "Mã CRM/CMS": c["ma_crm_cms"],
            "Tên KH": c["ten_kh"],
            "Bưu Cục": c["ma_bc_phu_trach"],
            "Phân Khúc": c["rfm_segment"],
            "Nhân Sự": c["nhan_su"] or "Chưa gán",
            "Kỳ Báo Cáo": c["current_rev"],
            "Kỳ So Sánh": c["previous_rev"],
            "Biến Động (VND)": c["diff_value"],
            "Tỷ Lệ (%)": c["diff_percent"],
            "Trạng Thái": "MỚI" if c["movement_status"] == "NEW" else ("RỜI BỎ" if c["movement_status"] == "CHURN" else ("TĂNG TRƯỞNG" if c["diff_value"] > 0 else "SỤT GIẢM"))
        })
        
    columns = [
        "STT", "Mã CRM/CMS", "Tên KH", "Bưu Cục", "Phân Khúc", 
        "Nhân Sự", "Kỳ Báo Cáo", "Kỳ So Sánh", "Biến Động (VND)",
        "Tỷ Lệ (%)", "Trạng Thái"
    ]
    df = pd.DataFrame(data, columns=columns)
    
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="BienDongKhachHang")
        
        workbook = writer.book
        worksheet = writer.sheets["BienDongKhachHang"]
        
        vnpost_blue = "0054A6"
        danger_red = "DC2626"
        success_green = "16A34A"
        
        header_fill = PatternFill(start_color=vnpost_blue, end_color=vnpost_blue, fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        alternate_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
        
        for col_idx, cell in enumerate(worksheet[1], 1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
            
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=len(df)+1), 2):
            is_alternate = row_idx % 2 == 0
            diff_val = row[8].value
            status_val = row[10].value
            
            for col_idx, cell in enumerate(row, 1):
                cell.border = thin_border
                cell.font = Font(size=10)
                
                if is_alternate:
                    cell.fill = alternate_fill
                
                if col_idx in [1, 2, 4, 5, 6]:
                    cell.alignment = center_align
                elif col_idx in [7, 8, 9, 10]:
                    cell.alignment = right_align
                    if col_idx in [7, 8, 9]:
                        cell.number_format = '#,##0 "đ"'
                    else:
                        cell.number_format = '0.0"%"'
                    if col_idx == 9:
                        if diff_val and diff_val > 0:
                            cell.font = Font(color=success_green, bold=True)
                        elif diff_val and diff_val < 0:
                            cell.font = Font(color=danger_red, bold=True)
                elif col_idx == 11:
                    cell.alignment = center_align
                    if status_val in ["MỚI", "TĂNG TRƯỞNG"]:
                        cell.font = Font(color=success_green, bold=True)
                    elif status_val in ["RỜI BỎ", "SỤT GIẢM"]:
                        cell.font = Font(color=danger_red, bold=True)
                else:
                    cell.alignment = left_align
        
        column_widths = {
            1: 5,   # STT
            2: 15,  # Mã
            3: 35,  # Tên KH
            4: 20,  # Bưu cục
            5: 15,  # Phân Khúc
            6: 20,  # Nhân sự
            7: 18,  # Kỳ Báo Cáo
            8: 18,  # Kỳ So Sánh
            9: 18,  # Biến Động
            10: 10, # Tỷ lệ
            11: 15  # Trạng thái
        }
        for col_idx, width in column_widths.items():
            column_letter = worksheet.cell(row=1, column=col_idx).column_letter
            worksheet.column_dimensions[column_letter].width = width
            
        worksheet.freeze_panes = "A2"

    buffer.seek(0)
    
    filename = "BaoCao_BienDongKH.xlsx"
    headers = {'Content-Disposition': f'attachment; filename="{filename}"'}
    return StreamingResponse(buffer, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
